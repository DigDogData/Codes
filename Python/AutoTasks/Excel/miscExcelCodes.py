#!/usr/bin/env python3

# miscExcelCodes.py - Example codes to interact with Excel

import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.utils import column_index_from_string
from openpyxl.styles import Font

wb = openpyxl.load_workbook("../Data/example.xlsx")
print(wb.sheetnames)

sheet = wb.active  # active sheet
print(sheet)
print(sheet["A1"].value)

anotherSheet = wb["Sheet3"]
print(anotherSheet)

c = sheet["B1"]  # get a cell from the sheet
print(c.value)
print("Row %s, Column %s is %s" % (c.row, c.column, c.value))
print("Cell %s is %s" % (c.coordinate, c.value))
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):  # go through every other row
    print(i, sheet.cell(row=i, column=2).value)

print("(", sheet.max_row, ",", sheet.max_column, ")")  # get size of the sheet

# convert column index to letter
print(get_column_letter(1))
print(get_column_letter(2))
print(get_column_letter(27))
print(get_column_letter(900))
print(get_column_letter(sheet.max_column))

# convert column letter to index
print(column_index_from_string("A"))
print(column_index_from_string("ABC"))

# get values of all cells between rows and columns
print(tuple(sheet["A1":"C3"]))  # get all cells from A1 to C3

for rowOfCellObjects in sheet["A1":"C3"]:
    for cellObj in rowOfCellObjects:
        print(cellObj.coordinate, cellObj.value)
    print("--- END OF ROW ---")

# get values of cells in a particular column (or row)
for cellObj in list(sheet.columns)[1]:
    print(cellObj.value)
print("---------------------------------")

# create and save excel documents
wb = openpyxl.Workbook()  # create a blank workbook
print(wb.sheetnames)  # starts with one sheet
sheet = wb.active
print(sheet.title)
sheet.title = "Spam Bacon Eggs"  # change sheet name/title
print(wb.sheetnames)
wb.create_sheet()
print(wb.sheetnames)
wb.create_sheet(index=0, title="First Sheet")
print(wb.sheetnames)
del wb["Sheet"]  # delete last sheet
print(wb.sheetnames)
# write value to cell
sheet = wb["First Sheet"]
sheet["A1"] = "Hello World!"
print(sheet["A1"].value)

# set font style of cells
wb = openpyxl.Workbook()  # create a blank workbook
sheet = wb["Sheet"]
italic24Font = Font(size=24, italic=True)  # create a font
sheet["A3"].font = italic24Font  # apply font to A1 cell
sheet["A3"] = "Hello World!"
wb.save("../Data/styles.xlsx")
