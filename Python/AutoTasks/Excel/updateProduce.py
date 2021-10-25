#!/usr/bin/env python3

# updateProduce.py - Corrects costs in produce sales spreadsheet

import openpyxl


class UpdatePrice:

    # initializer
    def __init__(self, sheet, updates):
        self.sheet = sheet
        self.updates = updates

    # method to update price
    def update_price(self):

        # loop through rows and update price
        for row in range(2, self.sheet.max_row):  # skip 1st row
            produce = self.sheet.cell(row=row, column=1).value  # get produce name

            # if name in updates list, update its price
            if produce in self.updates:
                self.sheet.cell(row=row, column=2).value = self.updates[produce]

        # save updated workbook to a new file
        print("Saving updated workbook...")
        wb.save("../Data/updatedProduceSales.xlsx")

        return None


if __name__ == "__main__":
    print("Opening workbook...")
    wb = openpyxl.load_workbook("../Data/produceSales.xlsx")
    sheet = wb["Sheet"]  # read sheet
    updates = {"Garlic": 3.07, "Celery": 1.19, "Lemon": 1.27}  # updated price
    UpdatePrice(sheet, updates).update_price()  # run method
