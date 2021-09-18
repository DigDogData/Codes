#!/usr/bin/env python3

# miscExcelCodes.py - Example codes to interact with Excel

import openpyxl

wb = openpyxl.load_workbook("example.xlsx")
print(wb.sheetnames)
sheet = wb.active  # active sheet
print(sheet)
print(sheet["A1"].value)
anotherSheet = wb["Sheet3"]
print(anotherSheet)
c = sheet["B1"]  # get a cell from the sheet
print(c.value)
print("Row %s, Column %s is %s" % (c.row, c.column, c.value))
print("Cell %s is %s" % (c.coordinate, c.value))
print(sheet.cell(row=1, column=2))
print(sheet.cell(row=1, column=2).value)

for i in range(1, 8, 2):  # go through every other row
    print(i, sheet.cell(row=i, column=2).value)

print("(", sheet.max_row, ",", sheet.max_column, ")")  # get size of the sheet
